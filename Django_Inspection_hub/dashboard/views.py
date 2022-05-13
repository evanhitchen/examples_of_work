from django.shortcuts import render, redirect
from django.views import generic
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.core.exceptions import ValidationError
from django.forms import inlineformset_factory, modelformset_factory
from django.conf import settings
import json
import openpyexcel
import openpyxl

# Create your views here.
from .models import *
from .forms import *
from .r9r9scripts import *
from .filters import *


def index(request):

    # Reference via html
    # <a href="{% url 'index' %}">Home</a>

    # Map data
    data = {'type': 'geojson',
            'data': {'type': 'FeatureCollection',
                     'features': [
                     ]}}

    sites = Site.objects.all()
    for site in sites:
        if site.latitude is None or site.longitude is None:
            pass
        else:
            site_dict = {'type': 'Feature',
                         'properties': {'description': f'{site.name}: '
                                                       f'{site.structure_set.all().count()} Structures'},
                         'geometry': {
                             'type': 'Point',
                             'coordinates': [site.longitude, site.latitude]}
                         }
            data['data']['features'].append(site_dict)

    js_data = json.dumps(data)

    # View function for home page of site.
    info = {}
    list_of_info = []

    clients = Client.objects.all()
    client_filter = dashboard_filter(request.GET, queryset=clients)
    clients = client_filter.qs

    for client in clients:
        sites = client.site_set.all()
        for site in sites:
            structures = site.structure_set.all()
            structure_filter1 = structure_filter(request.GET, queryset=structures)
            structures = structure_filter1.qs
            for structure in structures:
                defect_count = 0
                elements = structure.element_set.all()
                for element in elements:
                    defect_count += element.defect_set.all().count()
                list_of_info.append((client.name,
                                     site.name,
                                     structure.id_code,
                                     structure.name,
                                     structure.date_of_inspection,
                                     defect_count))

    context = {'table_info': list_of_info,
               'total_clients': Client.objects.all().count(),
               'total_sites': Site.objects.all().count(),
               'total_structures': Structure.objects.all().count(),
               'client_filter':client_filter,
               'structure_filter':structure_filter,
               'data': js_data}

    # Render the HTML template index.html with the data in the context variable
    return render(request, 'dashboard/index.html', context=context)

def client_info(request, pk):

    client = Client.objects.get(name=pk)

    list_of_info = []

    structure_count = 0

    sites = client.site_set.all()
    for site in sites:
        structures = site.structure_set.all()
        structure_count += site.structure_set.all().count()
        element_count = 0
        for structure in structures:
            defect_count = 0
            elements = structure.element_set.all()
            element_count += structure.element_set.all().count()
            for element in elements:
                defect_count += element.defect_set.all().count()
            list_of_info.append((site.name,
                                 structure.id_code,
                                 structure.name,
                                 structure.date_of_inspection,
                                 element_count,
                                 defect_count))

    context = {'table_info': list_of_info,
               'client_name': client.name,
               'total_sites': sites.count(),
               'total_structures': structure_count}

    return render(request, 'dashboard/client_info.html', context=context)


def site_info(request, pk):

    site = Site.objects.get(name=pk)
    list_of_info = []

    client = site.client

    structures = site.structure_set.all()
    element_count = 0
    for structure in structures:
        defect_count = 0
        elements = structure.element_set.all()
        element_count += structure.element_set.all().count()
        for element in elements:
            defect_count += element.defect_set.all().count()
        list_of_info.append((structure.id_code,
                             structure.name,
                             structure.date_of_inspection,
                             element_count,
                             defect_count))

    context = {'table_info': list_of_info,
               'client_name': client.name,
               'site_name': site.name}

    return render(request, 'dashboard/site_info.html', context=context)


def create_client(request):

    form = ClientForm()
    if request.method == 'POST':
        # print('printing post:', request.POST)
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_client_form.html', context=context)


def update_client(request, pk_test):

    client = Client.objects.get(name=pk_test)

    form = ClientForm(instance=client)

    if request.method == 'POST':
        # print('printing post:', request.POST)
        # print(request.POST.get('name'))
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_client_form.html', context=context)


def create_site(request):

    form = SiteForm()
    if request.method == 'POST':
        form = SiteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_site_form.html', context=context)


def update_site(request, pk_test):

    site = Site.objects.get(name=pk_test)

    form = SiteForm(instance=site)

    if request.method == 'POST':
        form = SiteForm(request.POST, instance=site)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_site_form.html', context=context)


def create_structure(request):

    form = StructureForm()
    if request.method == 'POST':
        form = StructureForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_structure_form.html', context=context)


def update_structure(request, pk_test):

    structure = Structure.objects.get(name=pk_test)

    form = StructureForm(instance=structure)

    if request.method == 'POST':
        form = SiteForm(request.POST, instance=structure)
        if form.is_valid():
            form.save()
            return redirect('/')

    context = {'form': form}
    return render(request, 'dashboard/new_structure_form.html', context=context)


def select_client(request):

    clients = Client_select_Form()

    if request.method == 'POST':
        client = request.POST.get('client_name')
        client_chosen = Client.objects.get(id=client)
        return redirect(f'/dashboard/update_client/{client_chosen}/')

    context = {'form': clients}
    return render(request, 'dashboard/select_client.html', context=context)


def upload_r8r9(request):
    form = DocxForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        form = DocxForm()
        # get object but keep it activated false so we only reference one variable and can call it below
        doc = Docxfile.objects.get(activated=False)
        doc.activated = True
        doc.save()
        try:
            output = fromr8r9(doc.file_name)
            client_name = output['owner']
            obj, created = Client.objects.get_or_create(name=client_name)
            if obj:
                client = obj
            else:
                client = created
            site_name = output['establishment_name']
            obj, created = Site.objects.get_or_create(name=site_name,
                                                      location=site_name,
                                                      client=client)
            if obj:
                site = obj
            else:
                site = created
            structure_id = output['ref_no']
            structure_name = output['structure_type']
            obj, created = Structure.objects.get_or_create(id_code=structure_id,
                                                           name=structure_name,
                                                           site=site)
            if obj:
                structure = obj
            else:
                structure = created
            for key in output['urgent_repairs']:
                element_info = output['urgent_repairs'][key]['Element']
                defect_info = output['urgent_repairs'][key]['Defect']
                obj, created = Element.objects.get_or_create(element=element_info,
                                                             element_description=element_info,
                                                             structure=structure)
                if obj:
                    element = obj
                else:
                    element = created
                obj, created = Defect.objects.get_or_create(defect_description=element_info,
                                                            remedial=' ',
                                                            safe='Y',
                                                            element=element)
                if obj:
                    defect = obj
                else:
                    defect = created
            for key in output['routine_repairs']:
                element_info = output['routine_repairs'][key]['Element']
                defect_info = output['routine_repairs'][key]['Defect']
                obj, created = Element.objects.get_or_create(element=element_info,
                                                             element_description=element_info,
                                                             structure=structure)
                if obj:
                    element = obj
                else:
                    element = created
                obj, created = Defect.objects.get_or_create(defect_description=element_info,
                                                            remedial=' ',
                                                            safe='Y',
                                                            element=element)
                if obj:
                    defect = obj
                else:
                    defect = created
            for key in output['preventative_maintenance_repairs']:
                element_info = output['preventative_maintenance_repairs'][key]['Element']
                defect_info = output['preventative_maintenance_repairs'][key]['Defect']
                obj, created = Element.objects.get_or_create(element=element_info,
                                                             element_description=element_info,
                                                             structure=structure)
                if obj:
                    element = obj
                else:
                    element = created
                obj, created = Defect.objects.get_or_create(defect_description=element_info,
                                                            remedial=' ',
                                                            safe='Y',
                                                            element=element)
                if obj:
                    defect = obj
                else:
                    defect = created
        except:
            raise ValidationError('Please check format of document to ensure it complies')

    context = {'form': form}
    return render(request,'dashboard/uploadr8r9.html', context=context)


def upload_excel(request):
    form = ExcelForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        form = ExcelForm()
        # get object but keep it activated false so we only reference one variable and can call it below
        doc = ExcelFile.objects.get(activated=False)
        doc.activated = True
        doc.save()
        try:
            workbook = openpyxl.load_workbook(filename=doc.file_name, data_only=False)
            sheet = workbook['Sheet1']
            row_count = sheet.max_row
            col_count = sheet.max_column
            structure_end_col = 8
            no_element_cols = 2
            no_of_defect_cols = 3
            data_dict = {}
            for i in range(int(row_count / 3)):
                data_dict[f'{i}'] = {}
                for j in range(structure_end_col):
                    key = sheet.cell(row=3 * i + 2, column=j + 1).value
                    data_dict[f'{i}'][key] = sheet.cell(row=3 * i + 3, column=j + 1).value
                    for k in range(structure_end_col, col_count):
                        if sheet.cell(row=3 * i + 2, column=k).value == 'element name':
                            element = sheet.cell(row=3 * i + 3, column=k).value
                            defect_counter = 1
                            for h in range(k, k + no_element_cols):
                                element_key = sheet.cell(row=3 * i + 2, column=h).value
                                data_dict[f'{i}']['elements'] = {}
                                data_dict[f'{i}']['elements'][element] = {}
                                data_dict[f'{i}']['elements'][element]['defects'] = {}
                                data_dict[f'{i}']['elements'][element][element_key] = sheet.cell(row=3 * i + 3,
                                                                                                 column=h).value
                            for g in range(k + no_element_cols, col_count):
                                if sheet.cell(row=3 * i + 2, column=g).value == 'element name':
                                    break
                                elif sheet.cell(row=3 * i + 2, column=g).value == 'defect description':
                                    data_dict[f'{i}']['elements'][element]['defects'][defect_counter] = {}
                                    for f in range(no_of_defect_cols):
                                        defect_key = sheet.cell(row=3 * i + 2, column=g + f).value
                                        data_dict[f'{i}']['elements'][element]['defects'][defect_counter][
                                            defect_key] = sheet.cell(row=3 * i + 3, column=g + f).value
                                    defect_counter += 1


            for key, value in data_dict.items():
                client_name = value['name']
                obj, created = Client.objects.get_or_create(name=client_name)
                if obj:
                    client = obj
                else:
                    client = created

                site_name = value['site name']
                location = value['location']
                latitude = value['latitude']
                longitude = value['longitude']
                obj, created = Site.objects.get_or_create(name=site_name,
                                                          location=location,
                                                          latitude=latitude,
                                                          longitude=longitude,
                                                          client=client)
                if obj:
                    site = obj
                else:
                    site = created

                structure_id = value['ID']
                structure_name = value['structure name']
                if value['date'] is None:
                    obj, created = Structure.objects.get_or_create(id_code=structure_id,
                                                                   name=structure_name,
                                                                   site=site)
                else:
                    date = value['date']
                    obj, created = Structure.objects.get_or_create(id_code=structure_id,
                                                                   name=structure_name,
                                                                   site=site,
                                                                   date_of_inspection=date)
                if obj:
                    structure = obj
                else:
                    structure = created
                for key2, value2 in value['elements'].items():
                    element_name = key2
                    element_description = value2['description']
                    obj, created = Element.objects.get_or_create(element=element_name,
                                                                 element_description=element_description,
                                                                 structure=structure)
                    if obj:
                        element = obj
                    else:
                        element = created

                    for key3, value3 in value2['defects'].items():
                        defect_description = value3['defect description']
                        remedial = value3['remedial']
                        safe = value3['safe']
                        obj, created = Defect.objects.get_or_create(defect_description=defect_description,
                                                                    remedial=remedial,
                                                                    safe=safe,
                                                                    element=element)
        except:
            raise ValidationError('Please check format of document to ensure it complies')

    context = {'form': form}
    return render(request, 'dashboard/uploadexcel.html', context=context)


def map(request):

    data = {'type': 'geojson',
            'data': {'type': 'FeatureCollection',
                     'features': [
                                  ]}}

    sites = Site.objects.all()
    for site in sites:
        if site.latitude is None or site.longitude is None:
            pass
        else:
            site_dict = {'type': 'Feature',
                         'properties': {'description': site.name},
                         'geometry': {
                                        'type': 'Point',
                                        'coordinates': [site.longitude, site.latitude]}
                         }
            data['data']['features'].append(site_dict)

    js_data = json.dumps(data)
    return render(request, 'dashboard/map.html', {"data": js_data})


def select_structure_for_R8R9(request):

    structures = Structure_select_Form()

    if request.method == 'POST':
        structure = request.POST.get('id_code')
        structure_chosen = Structure.objects.get(id=structure)
        return redirect('view_r8r9', pk=structure_chosen)

    context = {'form': structures}
    return render(request, 'dashboard/select_structure.html', context=context)


def view_r8r9(request, pk):

    structure = Structure.objects.get(id_code=pk)
    structure_form = StructureR8R9Form(instance=structure)
    elements = structure.element_set.all()
    defect_dict = {}
    for element in elements:
        defect_dict[f'{element.id}'] = []
        for defect in element.defect_set.all():
            defect_dict[f'{element.id}'].append(defect)

    ElementInlineFormSet = inlineformset_factory(Structure,
                                                 Element,
                                                 fields=('element', 'element_description',
                                                                             'discipline'), can_delete=False, extra=0,
                                                 widgets={'discipline':
                                                              CustomSelect(attrs={'class': 'form-control'}),
                                                          'element': forms.TextInput(attrs={'class': 'form-control'}),
                                                          'element_description':
                                                              forms.TextInput(attrs= {'class': 'form-control'}),
                                                          })

    ImageInLineFormset = inlineformset_factory(Defect, ImageFile, fields=('file_name',), extra=3,
                                               widgets={'file_name': forms.FileInput(attrs=
                                                                                     {'class': 'form-control-file'}),
                                               })

    form_dictionary = {}
    counter = 1
    defect_counter = 0
    for element in elements:
        form_dictionary[f'{counter}'] = {'formset': {},
                                         'counter': counter}
        for defect in defect_dict[f'{element.id}']:
            form_dictionary[f'{counter}']['formset'][f'{defect}'] = DefectR8R9Form(instance=defect,
                                                                                   prefix=f'{defect_counter}')
            form_dictionary[f'{counter}']['formset']['images'] = ImageInLineFormset(instance=defect,
                                                                                    prefix=f'images_{defect_counter}')
            defect_counter += 1
        counter += 1

    docform = DocxForm(request.POST or None, request.FILES or None)
    context = {'structure_form': structure_form,
               'elements': ElementInlineFormSet(instance=structure),
               'data': dict(form_dictionary),
               'docxform': docform}

    if request.method == 'POST':
        if 'Save_and_Generate' in request.POST:
            defect_counter = 0
            structure_form = StructureR8R9Form(request.POST, instance=structure)
            if structure_form.is_valid():
                edited_structure = structure_form.save()
                formset = ElementInlineFormSet(request.POST, request.FILES, instance=edited_structure)
                for form in formset:
                    if form.is_valid():
                        edited_element = form.save()
                        for defect in defect_dict[f'{edited_element.id}']:
                            defectform = DefectR8R9Form(request.POST, instance=defect, prefix=f'{defect_counter}')
                            if defectform.is_valid():
                                edited_defect = defectform.save(commit=False)
                                edited_defect.element = edited_element
                                edited_defect.save()
                                formset2 = ImageInLineFormset(request.POST, request.FILES, instance=edited_defect,
                                                              prefix=f'images_{defect_counter}')
                                if formset2.is_valid():
                                    print('yes')
                                    formset2.save()
                                defect_counter += 1
#
    return render(request, 'dashboard/view_r8r9edit.html', context=context)

